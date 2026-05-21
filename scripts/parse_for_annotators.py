import json
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

INPUT_FILE = "/mnt/user-data/uploads/results_max.json"
OUTPUT_EXCEL = "/mnt/user-data/outputs/annotator_task.xlsx"

os.makedirs("/mnt/user-data/outputs", exist_ok=True)


def clean_text(text):
    if not text:
        return ""
    text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', text)
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    text = re.sub(r' {2,}', ' ', text)
    return text.strip()


def remove_meta_commentary(text):
    """Remove parenthetical spoiler text and trailing meta sections."""
    if not text:
        return ""
    text = re.sub(
        r'\([^)]*(?:ovdje|spoiler|hint|zabluda|greška|pogreška|netočno|strategija|obmana'
        r'|this is|note:|napomena|objašnjenje|explanation|wrong|incorrect|misconception'
        r'|hallucination|false|untrue)[^)]*\)',
        '', text, flags=re.IGNORECASE
    )
    text = re.sub(
        r'\n+(?:\*+\s*)?(?:Why it|Why this|Note:|Explanation:|Napomena:|Zašto)[^\n]*.*$',
        '', text, flags=re.IGNORECASE | re.DOTALL
    )
    text = re.sub(r'\n{3,}', '\n\n', text)
    return clean_text(text)


def normalize_tags(text):
    """Normalize tag variants to plain [LABEL] form."""
    # Handle bold-inside-brackets: [**CORRECT**] -> [CORRECT]
    text = re.sub(
        r'\[\*+\s*(CORRECT|OBVIOUS(?:\s+HALLUCINATION)?|STRATEGIC(?:\s+MISCONCEPTION)?)\s*\*+\]',
        lambda m: '[' + m.group(1).split()[0] + ']',
        text, flags=re.IGNORECASE
    )
    # Handle bold-outside-brackets: **[CORRECT]** -> [CORRECT]
    text = re.sub(r'\*+\s*(\[[^\]]+\])\s*\*+', r'\1', text)
    # Normalize OBVIOUS HALLUCINATION -> OBVIOUS, STRATEGIC MISCONCEPTION -> STRATEGIC
    text = re.sub(r'\[\s*OBVIOUS\s+HALLUCINATION\s*\]', '[OBVIOUS]', text, flags=re.IGNORECASE)
    text = re.sub(r'\[\s*STRATEGIC\s+MISCONCEPTION\s*\]', '[STRATEGIC]', text, flags=re.IGNORECASE)
    return text


def split_sections(text):
    """Parse CORRECT / OBVIOUS / STRATEGIC sections from response text."""
    sections = {"CORRECT": "", "OBVIOUS": "", "STRATEGIC": ""}

    text = normalize_tags(text)

    pattern = r'\[\s*(CORRECT|OBVIOUS|STRATEGIC)\s*\]'
    matches = list(re.finditer(pattern, text, re.IGNORECASE))

    if matches:
        for i, match in enumerate(matches):
            label = match.group(1).upper()
            start = match.end()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
            content = clean_text(text[start:end])
            if len(content) > 10 and not sections[label]:
                sections[label] = content
    else:
        # Fallback: LABEL: style
        pattern2 = r'(CORRECT|OBVIOUS|STRATEGIC)\s*:'
        matches2 = list(re.finditer(pattern2, text, re.IGNORECASE))
        for i, match in enumerate(matches2):
            label = match.group(1).upper()
            start = match.end()
            end = matches2[i + 1].start() if i + 1 < len(matches2) else len(text)
            content = clean_text(text[start:end])
            if len(content) > 10 and not sections[label]:
                sections[label] = content

    for key in sections:
        sections[key] = remove_meta_commentary(sections[key])

    return sections


def main():
    with open(INPUT_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if isinstance(data, dict):
        for key in ['responses', 'data', 'results', 'items']:
            if key in data:
                data = data[key]
                break

    wb = Workbook()
    ws = wb.active
    ws.title = "ANOTACIJA"

    headers = ['#', 'ID', 'Koncept', 'Model', 'Tekst odgovora', 'class', 'error_type', 'persuasiveness']
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = header_fill
    ws.row_dimensions[1].height = 20

    row_num = 2
    total_written = 0
    missing = []

    label_order = ["CORRECT", "OBVIOUS", "STRATEGIC"]
    suffix_map = {"CORRECT": "_1", "OBVIOUS": "_2", "STRATEGIC": "_3"}

    fill_a = PatternFill("solid", fgColor="EBF1DE")
    fill_b = PatternFill("solid", fgColor="FFFFFF")

    data_sorted = sorted(data, key=lambda x: (x.get('id', ''), x.get('model', '')))

    prev_concept = None
    concept_idx = -1

    for item in data_sorted:
        concept_id = item.get('id', '')
        model = item.get('model', '')
        response = item.get('response', '')

        if not response:
            continue

        if concept_id != prev_concept:
            concept_idx += 1
            prev_concept = concept_id

        sections = split_sections(response)
        row_fill = fill_a if concept_idx % 2 == 0 else fill_b
        model_short = model.replace(':', '_').replace('/', '_')

        for label in label_order:
            text = sections[label]
            rid = f"{concept_id}_{model_short}{suffix_map[label]}"

            if not text:
                missing.append(rid)
                text = f"[MISSING {label}]"

            ws.cell(row=row_num, column=1, value=row_num - 1).font = Font(name="Arial", size=10)
            ws.cell(row=row_num, column=2, value=rid).font = Font(name="Arial", size=10)
            ws.cell(row=row_num, column=3, value=concept_id).font = Font(name="Arial", size=10)
            ws.cell(row=row_num, column=4, value=model).font = Font(name="Arial", size=10)

            text_cell = ws.cell(row=row_num, column=5, value=text)
            text_cell.font = Font(name="Arial", size=10)
            text_cell.alignment = Alignment(wrap_text=True, vertical='top')

            for col in [6, 7, 8]:
                ws.cell(row=row_num, column=col).font = Font(name="Arial", size=10)

            for col in range(1, 9):
                ws.cell(row=row_num, column=col).fill = row_fill

            ws.row_dimensions[row_num].height = 60

            row_num += 1
            total_written += 1

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 90
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16

    ws.freeze_panes = "A2"
    wb.save(OUTPUT_EXCEL)

    print("=" * 55)
    print("GOTOVO")
    print("=" * 55)
    print(f"Ukupno redaka: {total_written}")
    print(f"Excel: {OUTPUT_EXCEL}")
    if missing:
        print(f"Missing ({len(missing)}): {missing}")
    else:
        print("Nema missing sekcija")


if __name__ == "__main__":
    main()