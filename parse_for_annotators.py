import json
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

INPUT_FILE = "responses/results_max.json"
OUTPUT_EXCEL = "output/annotator_task.xlsx"


# ================== CLEAN ==================
def clean_text(text):
    if not text:
        return ""
    text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', text)
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    text = re.sub(r' {2,}', ' ', text)
    return text.strip()


def clean_model_response(text):
    """Soft clean — ne briše sadržaj"""
    if not text:
        return ""
    text = re.sub(r'\*Why it.*', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return clean_text(text)


# ================== PARSER ==================
def split_sections(text):
    sections = {
        "CORRECT": "",
        "OBVIOUS": "",
        "STRATEGIC": ""
    }

    # 1️⃣ [LABEL]
    pattern = r'\[\s*(CORRECT|OBVIOUS|STRATEGIC)[^\]]*\]'
    matches = list(re.finditer(pattern, text, re.IGNORECASE))

    if matches:
        for i, match in enumerate(matches):
            label = match.group(1).upper()
            start = match.end()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(text)

            content = clean_text(text[start:end])

            if len(content) > 10:
                sections[label] = content

    else:
        # 2️⃣ LABEL:
        pattern2 = r'(CORRECT|OBVIOUS|STRATEGIC)\s*:'
        matches2 = list(re.finditer(pattern2, text, re.IGNORECASE))

        if matches2:
            for i, match in enumerate(matches2):
                label = match.group(1).upper()
                start = match.end()
                end = matches2[i + 1].start() if i + 1 < len(matches2) else len(text)

                content = clean_text(text[start:end])

                if len(content) > 10:
                    sections[label] = content

    return sections


# ================== MAIN ==================
def parse_responses():
    os.makedirs("output", exist_ok=True)

    with open(INPUT_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if isinstance(data, dict):
        for key in ['responses', 'data', 'results', 'items']:
            if key in data:
                data = data[key]
                break

    wb = Workbook()
    ws = wb.active
    ws.title = "ANOTACIJA"

    headers = ['#', 'ID', 'Koncept', 'Tekst odgovora', 'class', 'error_type', 'persuasiveness']
    ws.append(headers)

    # styling (minimalno)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2

    total_expected = 0
    total_written = 0

    for item in data:
        concept_id = item.get('id', '')
        response = item.get('response', '')

        if not response:
            continue

        total_expected += 3

        sections = split_sections(response)

        # soft clean samo za strategic
        sections["STRATEGIC"] = clean_model_response(sections["STRATEGIC"])

        parts = [
            (f"{concept_id}_1", sections["CORRECT"], "CORRECT"),
            (f"{concept_id}_2", sections["OBVIOUS"], "OBVIOUS"),
            (f"{concept_id}_3", sections["STRATEGIC"], "STRATEGIC"),
        ]

        for rid, text, label in parts:

            if not text:
                print(f"⚠️ Missing: {rid}")
                text = f"[MISSING {label}]"

            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=rid)
            ws.cell(row=row_num, column=3, value=concept_id)
            ws.cell(row=row_num, column=4, value=text)

            row_num += 1
            total_written += 1

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 90

    wb.save(OUTPUT_EXCEL)

    print("\n" + "="*50)
    print("✅ GOTOVO")
    print("="*50)
    print(f"📊 Očekivano: {total_expected}")
    print(f"📊 Stvarno zapisano: {total_written}")
    print(f"📁 Excel: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    parse_responses()